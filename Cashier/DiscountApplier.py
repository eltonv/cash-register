# the as truncates it
import openpyxl as xl
# imports a library from pypi
from openpyxl.chart import BarChart, Reference, Series


def process_workbook(filename, discount):
    wb = xl.load_workbook(filename)
    # access the sheet
    sheet = wb['Sheet1']
    # takes 3rd column value of every row and applies formula
    # creating new column of values
    for row in range(2, sheet.max_row + 1):  # range second value is not included, so + 1
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 1 - (discount * 0.01)
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    # grabs the values from a sheet with the specified range
    values_disc = Reference(sheet,
                            min_row=2,
                            max_row=sheet.max_row,
                            min_col=4,
                            max_col=4)
    series2 = Series(values_disc, title="Discounted")

    values_orig = Reference(sheet,
                            min_row=2,
                            max_row=sheet.max_row,
                            min_col=3,
                            max_col=3)
    series1 = Series(values_orig, title="Original")

    chart = BarChart()
    chart.append(series1)
    chart.append(series2)
    chart.title = "Discount Price Comparison"

    sheet.add_chart(chart, 'e2')

    wb.save(filename)


# accessing a cell
# cell = sheet.cell(1,1)
# cell.value

